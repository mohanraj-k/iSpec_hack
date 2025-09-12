"""
MDD Output Generator for comprehensive file creation
Creates enriched output files using openpyxl and json directly
"""

import os
import json
import csv
import logging
from typing import Dict, List, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class MDDOutputGenerator:
    """Generates enriched MDD output files using openpyxl and json"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def generate_csv_output(self, results: List[Dict[str, Any]], output_path: str) -> bool:
        """Generate CSV file with all MDD data"""
        try:
            if not results:
                self.logger.warning("No results to write to CSV")
                return False
            
            # Define headers for CSV output
            headers = [
                'Allocation',
                'Logic creation date',
                'READY_FOR_DEV',
                'JARVIS/NON-JARVIS',
                'Study ID',
                'Plan Name',
                'IDRP CHECK NAME',
                'DQ Name',
                'Spotfire',
                'WORKFLOW STATUS',
                'DQ config status',
                'AVAILABLE_IN_SDQ',
                'REFERENCE',
                'Source Study ID',
                'Source DQ Name',
                'Note for Dev',
                'DQ Type',
                'Pseudo Code',
                'DQ Check Type',
                'Data Category: IDRP Data Category Name',
                'DQ Description',
                'Check logic',
                'Standard Query text',
                'IDRP VISIT NAME',
                'Primary Dataset',
                'Primary Dataset Columns',
                'Primary Form Name',
                'Primary Visit Name',
                'Primary Dynamic Domain',
                'Primary Dynamic Columns',
                'Query Target',
                'DQ Config:Target_dataset_refname',
                'DQ Config: Target item refname (ITEMID of RAW variable)',
                'Relational Dataset_1',
                'Relational Dataset Columns_1',
                'Relational Form Name_1',
                'Relational Visit Name_1',
                'Relational Dynamic Domain_1',
                'Relational Dynamic Columns_1',
                'Relational Dataset_2',
                'Relational Dataset Columns_2',
                'Relational Form Name_2',
                'Relational Visit Name_2',
                'Relational Dynamic Domain_2',
                'Relational Dynamic Columns_2',
                'SSPID Flag',
                'Relational Dataset_3',
                'Relational Dataset Columns_3',
                'Relational Form Name_3',
                'Relational Visit Name_3',
                'Relational Dynamic Domain_3',
                'Relational Dynamic Columns_3',
                'Relational Dataset_4',
                'Relational Dataset Columns_4',
                'Relational Form Name_4',
                'Relational Visit Name_4',
                'Relational Dynamic Domain_4',
                'Relational Dynamic Columns_4',
                'Relational Dataset_5',
                'Relational Dataset Columns_5',
                'Relational Form Name_5',
                'Relational Visit Name_5',
                'Relational Dynamic Domain_5',
                'Relational Dynamic Columns_5',
                'Logic modified date',
                'Logic modification reason',
                'Clarification Category',
                'Saama Comments',
                'Abbvie Comment',
                'DQ Config:Target_section_refname',
                'SENT_FOR_CREATION',
                'Reference Sponsor',
                'Reference Check Description',
                'Reference Query Text',
                'Is Match Found',
                'Confidence Score',
                'Match Type',
                'Match Reason',
                'Match Explanation'
            ]
            
            # Define fallback aliases for columns whose names may differ in `result`
            alias_map = {

                'Primary Form Name': ['P_form_name'],
                'Relational Form Name_1' : ['R1_form_name'],
                'Relational Form Name_2' : ['R2_form_name'],
                'Relational Form Name_3' : ['R3_form_name'],
                'Relational Form Name_4' : ['R4_form_name'],
                'Relational Form Name_5' : ['R5_form_name'],
                'Primary Visit Name': ['P_visit_name'],
                'Relational Visit Name_1' : ['R1_visit_name'],
                'Relational Visit Name_2' : ['R2_visit_name'],
                'Relational Visit Name_3' : ['R3_visit_name'],
                'Relational Visit Name_4' : ['R4_visit_name'],
                'Relational Visit Name_5' : ['R5_visit_name'],
                'Primary Dataset': ['Domain', 'Primary Domain','P_Domain'],
                'Relational Dataset_1' : ['R1_Domain'],
                'Relational Dataset_2' : ['R2_Domain'],
                'Relational Dataset_3' : ['R3_Domain'],
                'Relational Dataset_4' : ['R4_Domain'],
                'Relational Dataset_5' : ['R5_Domain'],
                'Primary Dataset Columns': ['Primary Domain Variables (Pre-Conf)', 'Primary Domain Variables'],
                'Relational Dynamic Columns_1' : ['R1_Dynamic_Columns'],
                'Relational Dynamic Columns_2' : ['R2_Dynamic_Columns'],
                'Relational Dynamic Columns_3' : ['R3_Dynamic_Columns'],
                'Relational Dynamic Columns_4' : ['R4_Dynamic_Columns'],
                'Relational Dynamic Columns_5' : ['R5_Dynamic_Columns'],
                'Primary Dynamic Domain' : ['Domain', 'Primary Domain','P_Domain'],
                'Relational Dynamic Domain_1' :['R1_Domain'],
                'Relational Dynamic Domain_2' :['R2_Domain'],
                'Relational Dynamic Domain_3' :['R3_Domain'],
                'Relational Dynamic Domain_4' :['R4_Domain'],
                'Relational Dynamic Domain_5' :['R5_Domain'],
                'Query Target': ['Query Target (Pre-Conf)', 'query_target'],
                'Relational Dataset Columns_1': ['Relational Domain Variables (Pre-Conf)','R1_Domain_Variables'],
                'Relational Dataset Columns_2': ['Relational Domain Variables (Pre-Conf)','R2_Domain_Variables'],
                'Relational Dataset Columns_3': ['Relational Domain Variables (Pre-Conf)','R3_Domain_Variables'],
                'Relational Dataset Columns_4': ['Relational Domain Variables (Pre-Conf)','R4_Domain_Variables'],
                'Relational Dataset Columns_5': ['Relational Domain Variables (Pre-Conf)','R5_Domain_Variables'],
                'Match Explanation': ['match_explanation'],
                # 'Study ID' :[],
                # 'Plan Name' :[],
                'IDRP CHECK NAME' : ['DQ Name'],
                'Source Study ID' : ['Reference Study'],
                'Source DQ Name' : ['Reference Check Name'],
                # 'Check logic' : [],
                'Standard Query text' : ['Standard Query Text'],
                'Data Category: IDRP Data Category Name' : ['Category'],


            }
            
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=headers)
                writer.writeheader()
                
                for result in results:
                    row = {}
                    for header in headers:
                        value = result.get(header, '')
                        # If direct key not found or empty, try alias list
                        if (value == '' or value is None) and header in alias_map:
                            for alias in alias_map[header]:
                                if alias in result and result.get(alias) not in [None, '']:
                                    value = result.get(alias)
                                    break
                        # Handle numeric conversion
                        if isinstance(value, (int, float)):
                            row[header] = str(value)
                        else:
                            row[header] = str(value) if value is not None else ''
                    writer.writerow(row)
            
            self.logger.info(f"Generated CSV output: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error generating CSV output: {str(e)}")
            return False

    def generate_excel_output(self, results: List[Dict[str, Any]], output_path: str) -> bool:
        """Generate enriched Excel file with all MDD data"""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Enriched MDD Results"
            
            if not results:
                self.logger.warning("No results to write to Excel")
                return False
            
            # Define headers based on first result
            sample_result = results[0]
            headers = [
                'Target Check Description',
                'Target Query Text', 
                'Match Classification',
                'Confidence Score',
                'Reference Check Name',
                'Reference Description',
                'Reference Query Text',
                'Reference Form Name',
                'Reference Visit Name',
                'Match Explanation',
                'Is Match Found'
            ]
            
            # Write headers with formatting
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Write data rows
            for row_idx, result in enumerate(results, 2):
                # Target Check Description
                sheet.cell(row=row_idx, column=1, value=result.get('Target Check Description', ''))
                
                # Target Query Text
                sheet.cell(row=row_idx, column=2, value=result.get('Target Query Text', ''))
                
                # Match Classification
                classification = result.get('match_classification', 'No Match')
                cell = sheet.cell(row=row_idx, column=3, value=classification)
                
                # Color code based on classification
                if classification == 'Excellent':
                    cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                elif classification == 'Good':
                    cell.fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
                elif classification == 'Moderate':
                    cell.fill = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')
                elif classification == 'Weak':
                    cell.fill = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
                
                # Confidence Score
                confidence = result.get('confidence_score', 0)
                sheet.cell(row=row_idx, column=4, value=f"{confidence:.3f}" if confidence else "0.000")
                
                # Reference fields
                sheet.cell(row=row_idx, column=5, value=result.get('DQ Name', ''))
                sheet.cell(row=row_idx, column=6, value=result.get('DQ description', ''))
                sheet.cell(row=row_idx, column=7, value=result.get('Standard Query text', ''))
                sheet.cell(row=row_idx, column=8, value=result.get('Primary Form Name', ''))
                sheet.cell(row=row_idx, column=9, value=result.get('Primary Visit Name', ''))
                
                # Match Explanation
                explanation = result.get('match_explanation', result.get('match_reason', ''))
                sheet.cell(row=row_idx, column=10, value=explanation)
                
                # Is Match Found
                sheet.cell(row=row_idx, column=11, value=result.get('is_match_found', 'NO'))
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            workbook.save(output_path)
            workbook.close()
            
            self.logger.info(f"Generated Excel output: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error generating Excel output: {e}")
            return False

    def generate_json_output(self, results: List[Dict[str, Any]], output_path: str, metadata: Dict[str, Any] = None) -> bool:
        """Generate detailed JSON output with metadata"""
        try:
            # Prepare output structure
            output_data = {
                'metadata': {
                    'generated_at': datetime.now().isoformat(),
                    'total_results': len(results),
                    'algorithm_version': 'Enhanced Semantic Matching v2.0',
                    'scoring_method': '50% cosine similarity + 50% keyword overlap',
                    'classification_thresholds': {
                        'excellent': '≥35%',
                        'good': '≥25%', 
                        'moderate': '≥15%',
                        'weak': '≥5%'
                    }
                },
                'results': results
            }
            
            # Add custom metadata if provided
            if metadata:
                output_data['metadata'].update(metadata)
            
            # Calculate statistics
            if results:
                classifications = [r.get('match_classification', 'No Match') for r in results]
                stats = {
                    'excellent_matches': classifications.count('Excellent'),
                    'good_matches': classifications.count('Good'),
                    'moderate_matches': classifications.count('Moderate'), 
                    'weak_matches': classifications.count('Weak'),
                    'no_matches': classifications.count('No Match')
                }
                
                total_with_matches = sum([stats['excellent_matches'], stats['good_matches'], 
                                        stats['moderate_matches'], stats['weak_matches']])
                
                stats['match_rate'] = f"{(total_with_matches / len(results) * 100):.1f}%" if results else "0%"
                
                output_data['metadata']['statistics'] = stats
            
            # Write JSON file
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, indent=2, ensure_ascii=False)
            
            self.logger.info(f"Generated JSON output: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error generating JSON output: {e}")
            return False