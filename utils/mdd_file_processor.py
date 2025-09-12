"""
MDD file processor for comprehensive Excel file handling
Uses openpyxl directly for robust Excel file processing
"""

import os
import logging
import re
from typing import Dict, List, Any, Optional
import openpyxl
import pandas as pd
from datetime import datetime
from utils.field_aliases import get_field_value as _gf

class MDDFileProcessor:
    """Comprehensive processor for MDD Excel files using openpyxl"""
    
    def __init__(self):
        """Initialize file processor"""
        self.logger = logging.getLogger(__name__)

    def parse_target_mdd(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse target MDD file"""
        try:
            ext = os.path.splitext(file_path)[1].lower()
            target_data: List[Dict[str, Any]] = []

            if ext == '.csv':
                # Parse CSV using pandas
                df = pd.read_csv(file_path, dtype=str, keep_default_na=False, encoding='utf-8-sig',encoding_errors="replace")
                headers = [str(h) if h is not None else '' for h in df.columns]

                for i, row in df.iterrows():
                    row_dict: Dict[str, Any] = {}
                    for h in headers:
                        value = row.get(h, '')
                        row_dict[h] = str(value) if value is not None else ''

                    if self._is_valid_target_row(row_dict):
                        # Row index aligns with Excel convention: header at row 1, first data row at 2
                        row_dict['row_index'] = int(i) + 2
                        target_data.append(row_dict)

                self.logger.info(f"Parsed target MDD CSV file: {len(target_data)},{pd.DataFrame(target_data).shape} valid rows")
                return target_data
            else:
                # Default: parse Excel using openpyxl
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                sheet = workbook.active
                
                # Get headers from first row
                headers = []
                for cell in sheet[1]:
                    headers.append(str(cell.value) if cell.value else '')
                
                # Process data rows
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    row_dict = {}
                    for col_idx, value in enumerate(row):
                        if col_idx < len(headers):
                            row_dict[headers[col_idx]] = str(value) if value is not None else ''
                    
                    # Check if row has meaningful content
                    if self._is_valid_target_row(row_dict):
                        row_dict['row_index'] = row_idx
                        target_data.append(row_dict)
                
                workbook.close()
                self.logger.info(f"Parsed target MDD file: {len(target_data)},{pd.DataFrame(target_data).shape} valid rows")
                return target_data
            
        except Exception as e:
            self.logger.error(f"Error parsing target MDD file {file_path}: {e}")
            return []

    def clean_row_columns(self, row_dict: dict, headers: list) -> dict:
        """Normalize column values in a row dict.

        - Strip [] and quotes from list-like strings for known columns.
        - Additionally, for any dynamic_columns* fields: if the value looks like
          a dict-like string such as "{KEY1: value1, KEY2: value2}", extract only
          the keys and store as a comma-separated string: "KEY1, KEY2".
        """
        # self.logger.info(f"Rs-line54 - logging row_dict: {row_dict}")
        columns_to_clean = [
            'primary_dataset_columns', 'primary_form_name', 'primary_visit_name', 'primary_dynamic_domain', 'primary_dynamic_columns',
            'relational_dataset_1', 'relational_dataset_columns_1', 'relational_form_name_1', 'relational_visit_name_1', 'relational_dynamic_domain_1', 'relational_dynamic_columns_1',
            'relational_dataset_2', 'relational_dataset_columns_2', 'relational_form_name_2', 'relational_visit_name_2', 'relational_dynamic_domain_2', 'relational_dynamic_columns_2',
            'relational_dataset_3', 'relational_dataset_columns_3', 'relational_form_name_3', 'relational_visit_name_3', 'relational_dynamic_domain_3', 'relational_dynamic_columns_3',
            'relational_dataset_4', 'relational_dataset_columns_4', 'relational_form_name_4', 'relational_visit_name_4', 'relational_dynamic_domain_4', 'relational_dynamic_columns_4',
            'relational_dataset_5', 'relational_dataset_columns_5', 'relational_form_name_5', 'relational_visit_name_5', 'relational_dynamic_domain_5', 'relational_dynamic_columns_5',
        ]
        for col in columns_to_clean:
            if col in row_dict and isinstance(row_dict[col], str):
                s = row_dict[col].strip()
                # Handle dict-like strings specifically for dynamic_columns*
                if ('dynamic_columns' in col) and ('{' in s) and (':' in s):
                    try:
                        # Capture tokens before ':' inside the braces. This is robust even if values contain commas.
                        keys = re.findall(r'([^{},:]+)\s*:', s)
                        # Clean and deduplicate while preserving order
                        cleaned = []
                        seen = set()
                        for k in keys:
                            k = k.strip().strip("'\"")
                            if k and k not in seen:
                                seen.add(k)
                                cleaned.append(k)
                        row_dict[col] = ', '.join(cleaned)
                    except Exception:
                        # Fallback to the generic cleanup
                        row_dict[col] = s.replace('[', '').replace(']', '').replace('"', '').replace("'", '')
                else:
                    # Generic cleanup for list-like strings
                    row_dict[col] = s.replace('[', '').replace(']', '').replace('"', '').replace("'", '')
        return row_dict

    def parse_reference_mdd(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse reference MDD file"""
        try:
            ext = os.path.splitext(file_path)[1].lower()
            reference_data: List[Dict[str, Any]] = []
            # Prepare report extract date stamp if missing in input
            report_col = 'report_extract_date'
            now_dt = datetime.now()
            now_str = now_dt.strftime("%Y-%m-%d %H:%M:%S.") + f"{int(now_dt.microsecond / 1000):03d}"

            if ext == '.csv':
                # Parse CSV reference MDD using pandas
                df = pd.read_csv(file_path, dtype=str, keep_default_na=False, encoding='utf-8-sig')
                headers = [str(h) if h is not None else '' for h in df.columns]
                # Determine if report_extract_date column exists
                has_report_col = any(str(h).strip().lower() == report_col for h in headers)
                # self.logger.info(f"Rs-line127 - logging headers: {headers},{len(df)}")
                for i, row in df.iterrows():
                    row_dict: Dict[str, Any] = {}
                    for h in headers:
                        value = row.get(h, '')
                        row_dict[h] = str(value) if value is not None else ''

                    if not has_report_col:
                        row_dict[report_col] = now_str
                    if self._is_valid_reference_row(row_dict):
                        row_dict['row_index'] = int(i) + 2
                        row_dict['source_file'] = os.path.basename(file_path)
                        reference_data.append(row_dict)

                reference_data = [self.clean_row_columns(row, headers) for row in reference_data]
                self.logger.info(f"Parsed reference MDD CSV file: {len(reference_data)} {pd.DataFrame(reference_data).shape} valid rows")
                return reference_data
            else:
                # Default: parse Excel reference MDD using openpyxl
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                sheet = workbook.active
                # self.logger.info(f'Rs-line55 - logging sheet names: {workbook.sheetnames}')
                # Get headers from first row
                headers = []
                for cell in sheet[1]:
                    headers.append(str(cell.value) if cell.value else '')
                # Determine if report_extract_date column exists
                has_report_col = any(str(h).strip().lower() == report_col for h in headers)
                # self.logger.info(f'Rs-line60 - logging headers: {headers}')
                
                # Process data rows
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    row_dict = {}
                    # self.logger.info(f'Rs-line66 - logging row: {row_idx},{row}')
                    for col_idx, value in enumerate(row):
                        # self.logger.info(f'Rs-line68 - logging col_idx: {col_idx},{len(headers)}')
                        if col_idx < len(headers):
                            row_dict[headers[col_idx]] = str(value) if value is not None else ''
                            # self.logger.info(f'Rs-line71 - logging row_dict: {row_dict}')
                    if not has_report_col:
                        row_dict[report_col] = now_str
                    # Check if row has meaningful content
                    # self.logger.info(f'Rs-line73 - logging row_dict: {len(row_dict)},{self._is_valid_reference_row(row_dict)}')
                    if self._is_valid_reference_row(row_dict):
                        row_dict['row_index'] = row_idx
                        row_dict['source_file'] = os.path.basename(file_path)
                        reference_data.append(row_dict)
                        # self.logger.info(f'Rs-line78 - logging reference_data: {reference_data}')   
                
                workbook.close()
                reference_data = [self.clean_row_columns(row, headers) for row in reference_data]#[:10] ##REMOVE Slicing
                self.logger.info(f"Parsed reference MDD file: {len(reference_data)} {pd.DataFrame(reference_data).shape} valid rows")
                # self.logger.info(f'Rs-line82 - logging final reference_data: {reference_data}')
                return reference_data
            
        except Exception as e:
            self.logger.error(f"Error parsing reference MDD file {file_path}: {e}")
            return []

    # def _is_valid_target_row(self, row_dict: Dict[str, Any]) -> bool:
    #     """Check if target row is valid (has required fields)"""
    #     # Check for key fields that indicate a valid target row
    #     key_fields = ['DQ Name', 'DQ name', 'name', 'DQ Description', 'DQ description', 'description']
        
    #     for field in key_fields:
    #         value = row_dict.get(field, '')
    #         if value and str(value).strip() and str(value).strip().lower() not in ['n/a', 'na', '']:
    #             return True
        
    #     return False
    def _is_valid_target_row(self, row_dict: Dict[str, Any]) -> bool:
        """Check if target row contains all mandatory fields: DQ Name, DQ Description, Standard Query text"""
        # Require all mandatory fields using canonical alias mapping
        dq_name = _gf(row_dict, 'dq_name')
        dq_desc = _gf(row_dict, 'dq_description')
        query_text = _gf(row_dict, 'query_text')
        
        return bool(dq_name and dq_desc and query_text)

    def _is_valid_reference_row(self, row_dict: Dict[str, Any]) -> bool:
        """Check if reference row contains all mandatory fields: DQ Name, DQ Description, Standard Query text"""
        # Require all mandatory fields using canonical alias mapping
        dq_name = _gf(row_dict, 'dq_name')
        dq_desc = _gf(row_dict, 'dq_description')
        query_text = _gf(row_dict, 'query_text')
        
        return bool(dq_name and dq_desc and query_text)

    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """Get information about an Excel file"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            info = {
                'filename': os.path.basename(file_path),
                'sheet_name': sheet.title,
                'total_rows': sheet.max_row,
                'total_columns': sheet.max_column,
                'file_size': os.path.getsize(file_path)
            }
            
            workbook.close()
            return info
            
        except Exception as e:
            self.logger.error(f"Error getting file info for {file_path}: {e}")
            return {'error': str(e)}

    def validate_mdd_structure(self, file_path: str) -> Dict[str, Any]:
        """Validate MDD file structure"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            # Get headers
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value) if cell.value else '')
            
            # Check for expected MDD columns
            expected_columns = ['DQ Name', 'DQ Description', 'Standard Query text', 'Primary Form Name', 'Primary Visit Name']
            found_columns = []
            
            for expected in expected_columns:
                for header in headers:
                    if expected.lower() in header.lower():
                        found_columns.append(header)
                        break
            
            validation_result = {
                'valid': len(found_columns) >= 2,  # At least 2 key columns found
                'headers': headers,
                'found_columns': found_columns,
                'missing_columns': [col for col in expected_columns if col not in found_columns],
                'total_rows': sheet.max_row - 1  # Exclude header row
            }
            
            workbook.close()
            return validation_result
            
        except Exception as e:
            self.logger.error(f"Error validating MDD structure for {file_path}: {e}")
            return {'valid': False, 'error': str(e)}